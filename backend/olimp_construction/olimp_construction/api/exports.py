"""Экспорт КС-2 и КС-3 в PDF (через wkhtmltopdf) и Excel (через openpyxl).

PDF использует штатный Frappe printview с Print Format «КС-2/КС-3 (унифицированная форма)».
Excel генерируется напрямую по полям документа в формате, приближённом к
унифицированным формам Госкомстата (ОКУД 0322005 / 0322001).
"""
from __future__ import annotations

import io

import frappe
from frappe import _
from frappe.utils import flt, formatdate
from frappe.utils.pdf import get_pdf

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


KS2_PRINT_FORMAT = "КС-2 (унифицированная форма)"
KS3_PRINT_FORMAT = "КС-3 (унифицированная форма)"


def _fmt_money(value) -> str:
    if value is None:
        return ""
    return f"{flt(value):,.2f}".replace(",", " ").replace(".", ",")


def _fmt_qty(value) -> str:
    if value is None:
        return ""
    return f"{flt(value):,.3f}".replace(",", " ").replace(".", ",")


def _fmt_date(value) -> str:
    if not value:
        return ""
    return formatdate(value, "dd.MM.yyyy")


def _render_html(doctype: str, name: str, print_format: str) -> str:
    """Рендерит Print Format в HTML с подмешиванием project_title."""
    if not frappe.db.exists("Print Format", print_format):
        frappe.throw(_(f"Print Format «{print_format}» не найден. Запустите setup_print_formats()."))

    doc = frappe.get_doc(doctype, name)
    project_title = ""
    if doc.get("project") and frappe.db.exists("Construction Project", doc.project):
        project_title = frappe.db.get_value("Construction Project", doc.project, "title") or doc.project

    pf = frappe.get_doc("Print Format", print_format)
    template = pf.html or ""
    rendered = frappe.render_template(template, {
        "doc": doc,
        "project_title": project_title,
        "frappe": frappe,
    })

    # Оборачиваем в полный HTML-документ для wkhtmltopdf
    return f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8" />
<title>{print_format} — {doc.name}</title>
</head>
<body>{rendered}</body>
</html>
"""


def _pdf_response(html: str, filename: str) -> None:
    """Возвращает PDF клиенту через Frappe response."""
    pdf_bytes = get_pdf(html, options={
        "page-size": "A4",
        "orientation": "Landscape" if "0322005" in html or "КС-2" in filename else "Portrait",
        "margin-top": "10mm",
        "margin-bottom": "10mm",
        "margin-left": "10mm",
        "margin-right": "10mm",
        "encoding": "UTF-8",
    })

    frappe.local.response.filename = filename
    frappe.local.response.filecontent = pdf_bytes
    frappe.local.response.type = "pdf"


@frappe.whitelist()
def ks2_pdf(name: str) -> None:
    """Скачивание PDF КС-2 в гос.форме."""
    frappe.has_permission("KS2 Act", "read", doc=name, throw=True)
    html = _render_html("KS2 Act", name, KS2_PRINT_FORMAT)
    _pdf_response(html, f"KS-2_{name}.pdf")


@frappe.whitelist()
def ks3_pdf(name: str) -> None:
    """Скачивание PDF КС-3 в гос.форме."""
    frappe.has_permission("KS3 Act", "read", doc=name, throw=True)
    html = _render_html("KS3 Act", name, KS3_PRINT_FORMAT)
    _pdf_response(html, f"KS-3_{name}.pdf")


# ------------------------- Excel ------------------------------------

_THIN = Side(border_style="thin", color="000000")
_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="F0F0F0")
_TOTAL_FILL = PatternFill("solid", fgColor="F8F8F8")
_FONT_HDR = Font(name="Times New Roman", size=10, bold=True)
_FONT_BODY = Font(name="Times New Roman", size=10)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


def _excel_response(wb: Workbook, filename: str) -> None:
    buf = io.BytesIO()
    wb.save(buf)
    frappe.local.response.filename = filename
    frappe.local.response.filecontent = buf.getvalue()
    frappe.local.response.type = "binary"
    frappe.local.response.headers = {
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }


def _apply_border(ws, cell_range: str) -> None:
    for row in ws[cell_range]:
        for cell in row:
            cell.border = _BORDER_ALL


@frappe.whitelist()
def ks2_excel(name: str) -> None:
    """Excel-выгрузка КС-2 в формате, приближённом к ОКУД 0322005."""
    frappe.has_permission("KS2 Act", "read", doc=name, throw=True)
    doc = frappe.get_doc("KS2 Act", name)

    project_title = ""
    if doc.project and frappe.db.exists("Construction Project", doc.project):
        project_title = frappe.db.get_value("Construction Project", doc.project, "title") or doc.project

    wb = Workbook()
    ws = wb.active
    ws.title = "КС-2"

    # Ширина колонок (8 колонок госформы)
    widths = [5, 38, 14, 8, 11, 14, 16, 6]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    # Шапка с реквизитами
    meta = [
        ("Инвестор:", ""),
        ("Заказчик (Генподрядчик):", doc.customer or ""),
        ("Подрядчик (Субподрядчик):", "ООО «Олимп», г. Екатеринбург"),
        ("Стройка:", project_title),
        ("Объект:", doc.title or ""),
        ("Договор подряда (контракт):", f"№ {doc.contract_number or '—'}"),
    ]
    for label, value in meta:
        ws.cell(row=row, column=1, value=label).font = Font(name="Times New Roman", size=9, italic=True)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        ws.cell(row=row, column=2, value=value).font = _FONT_BODY
        ws.cell(row=row, column=7, value="Форма по ОКУД" if row == 1 else "").font = Font(size=8)
        ws.cell(row=row, column=8, value="0322005" if row == 1 else "").font = _FONT_HDR
        ws.cell(row=row, column=8).alignment = _ALIGN_CENTER
        if row == 1:
            ws.cell(row=row, column=8).border = _BORDER_ALL
        row += 1

    row += 1
    # Номер / Дата / Период
    ws.cell(row=row, column=1, value="Номер документа").alignment = _ALIGN_CENTER
    ws.cell(row=row, column=1).font = Font(size=8)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws.cell(row=row, column=2, value="Дата составления").alignment = _ALIGN_CENTER
    ws.cell(row=row, column=2).font = Font(size=8)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    ws.cell(row=row, column=4, value="Отчётный период").alignment = _ALIGN_CENTER
    ws.cell(row=row, column=4).font = Font(size=8)
    row += 1
    ws.cell(row=row, column=1, value=doc.act_number or doc.name).alignment = _ALIGN_CENTER
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws.cell(row=row, column=2, value=_fmt_date(doc.act_date)).alignment = _ALIGN_CENTER
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    ws.cell(row=row, column=4,
            value=f"с {_fmt_date(doc.period_from)} по {_fmt_date(doc.period_to)}").alignment = _ALIGN_CENTER
    _apply_border(ws, f"A{row-1}:F{row}")
    row += 2

    # Заголовок
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value="АКТ о приёмке выполненных работ")
    cell.font = Font(name="Times New Roman", size=14, bold=True)
    cell.alignment = _ALIGN_CENTER
    row += 2

    # Шапка таблицы (8 колонок)
    headers = [
        "№ п/п",
        "Наименование работ",
        "Номер единичной расценки",
        "Единица измерения",
        "Количество",
        "Цена за единицу, руб.",
        "Стоимость, руб.",
        "Номер пункта",
    ]
    table_start = row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = _FONT_HDR
        cell.alignment = _ALIGN_CENTER
        cell.fill = _HEADER_FILL
        cell.border = _BORDER_ALL
    ws.row_dimensions[row].height = 36
    row += 1
    # Номера колонок
    for col in range(1, 9):
        cell = ws.cell(row=row, column=col, value=col)
        cell.alignment = _ALIGN_CENTER
        cell.font = Font(name="Times New Roman", size=8, italic=True)
        cell.border = _BORDER_ALL
    row += 1

    # Позиции
    for idx, item in enumerate(doc.items or [], 1):
        ws.cell(row=row, column=1, value=idx).alignment = _ALIGN_CENTER
        ws.cell(row=row, column=2, value=item.work_name or "").alignment = _ALIGN_LEFT
        ws.cell(row=row, column=3, value=item.estimate_ref or "").alignment = _ALIGN_CENTER
        ws.cell(row=row, column=4, value=item.unit or "").alignment = _ALIGN_CENTER
        ws.cell(row=row, column=5, value=flt(item.qty) if item.qty else None).alignment = _ALIGN_RIGHT
        ws.cell(row=row, column=5).number_format = '#,##0.000'
        ws.cell(row=row, column=6, value=flt(item.unit_price) if item.unit_price else None).alignment = _ALIGN_RIGHT
        ws.cell(row=row, column=6).number_format = '#,##0.00'
        ws.cell(row=row, column=7, value=flt(item.amount) if item.amount else None).alignment = _ALIGN_RIGHT
        ws.cell(row=row, column=7).number_format = '#,##0.00'
        ws.cell(row=row, column=8, value="").alignment = _ALIGN_CENTER
        for col in range(1, 9):
            ws.cell(row=row, column=col).font = _FONT_BODY
            ws.cell(row=row, column=col).border = _BORDER_ALL
        row += 1

    # Итого
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value="Итого:").alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=1).font = _FONT_HDR
    ws.cell(row=row, column=7, value=flt(doc.amount)).number_format = '#,##0.00'
    ws.cell(row=row, column=7).alignment = _ALIGN_RIGHT
    ws.cell(row=row, column=7).font = _FONT_HDR
    for col in range(1, 9):
        ws.cell(row=row, column=col).fill = _TOTAL_FILL
        ws.cell(row=row, column=col).border = _BORDER_ALL
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value="Всего по акту:").alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=1).font = _FONT_HDR
    ws.cell(row=row, column=7, value=flt(doc.amount)).number_format = '#,##0.00'
    ws.cell(row=row, column=7).alignment = _ALIGN_RIGHT
    ws.cell(row=row, column=7).font = _FONT_HDR
    for col in range(1, 9):
        ws.cell(row=row, column=col).fill = _TOTAL_FILL
        ws.cell(row=row, column=col).border = _BORDER_ALL
    row += 3

    # Подписи
    ws.cell(row=row, column=1, value="Сдал (Подрядчик):").font = _FONT_HDR
    ws.cell(row=row, column=5, value="Принял (Заказчик):").font = _FONT_HDR
    row += 2
    ws.cell(row=row, column=1, value="Директор ООО «Олимп»  _______________  Клочков Д.А.").font = _FONT_BODY
    ws.cell(row=row, column=5, value="_______________  _______________").font = _FONT_BODY
    row += 1
    ws.cell(row=row, column=1, value="должность / подпись / расшифровка").font = Font(size=8, italic=True)
    ws.cell(row=row, column=5, value="должность / подпись / расшифровка").font = Font(size=8, italic=True)
    row += 1
    ws.cell(row=row, column=1, value="М.П.").font = Font(size=9)
    ws.cell(row=row, column=5, value="М.П.").font = Font(size=9)

    _excel_response(wb, f"KS-2_{doc.name}.xlsx")


@frappe.whitelist()
def ks3_excel(name: str) -> None:
    """Excel-выгрузка КС-3 в формате, приближённом к ОКУД 0322001."""
    frappe.has_permission("KS3 Act", "read", doc=name, throw=True)
    doc = frappe.get_doc("KS3 Act", name)

    project_title = ""
    if doc.project and frappe.db.exists("Construction Project", doc.project):
        project_title = frappe.db.get_value("Construction Project", doc.project, "title") or doc.project

    wb = Workbook()
    ws = wb.active
    ws.title = "КС-3"

    # Ширина 6 колонок госформы
    widths = [5, 42, 10, 16, 16, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    meta = [
        ("Инвестор:", ""),
        ("Заказчик (Генподрядчик):",
         (doc.customer or "") + (f", ИНН {doc.customer_inn}" if doc.customer_inn else "")),
        ("Адрес заказчика:", doc.customer_address or ""),
        ("Подрядчик (Субподрядчик):",
         "ООО «Олимп»" + (f", ИНН {doc.contractor_inn}" if doc.contractor_inn else "")),
        ("Адрес подрядчика:", doc.contractor_address or "г. Екатеринбург"),
        ("Стройка:", project_title),
        ("Договор подряда:", f"№ {doc.contract_number or '—'}"),
    ]
    for label, value in meta:
        ws.cell(row=row, column=1, value=label).font = Font(name="Times New Roman", size=9, italic=True)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row=row, column=2, value=value).font = _FONT_BODY
        ws.cell(row=row, column=5, value="Форма по ОКУД" if row == 1 else "").font = Font(size=8)
        ws.cell(row=row, column=6, value=doc.okud_code or "0322001" if row == 1 else "").font = _FONT_HDR
        ws.cell(row=row, column=6).alignment = _ALIGN_CENTER
        if row == 1:
            ws.cell(row=row, column=6).border = _BORDER_ALL
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Номер документа").font = Font(size=8)
    ws.cell(row=row, column=1).alignment = _ALIGN_CENTER
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws.cell(row=row, column=2, value="Дата составления").font = Font(size=8)
    ws.cell(row=row, column=2).alignment = _ALIGN_CENTER
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    ws.cell(row=row, column=4, value="Отчётный период").font = Font(size=8)
    ws.cell(row=row, column=4).alignment = _ALIGN_CENTER
    row += 1
    ws.cell(row=row, column=1, value=doc.act_number or doc.name).alignment = _ALIGN_CENTER
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws.cell(row=row, column=2, value=_fmt_date(doc.report_date)).alignment = _ALIGN_CENTER
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    ws.cell(row=row, column=4,
            value=f"с {_fmt_date(doc.period_from)} по {_fmt_date(doc.period_to)}").alignment = _ALIGN_CENTER
    _apply_border(ws, f"A{row-1}:F{row}")
    row += 2

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value="СПРАВКА о стоимости выполненных работ и затрат")
    cell.font = Font(name="Times New Roman", size=14, bold=True)
    cell.alignment = _ALIGN_CENTER
    row += 2

    headers_top = [
        ("№ п/п", 1, 1),
        ("Наименование пусковых комплексов, объектов,\nвидов работ, оборудования, затрат", 1, 1),
        ("Код", 1, 1),
        ("Стоимость выполненных работ и затрат, руб.", 1, 3),
    ]
    col = 1
    for text, _rs, cs in headers_top:
        if cs > 1:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + cs - 1)
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = _FONT_HDR
        cell.fill = _HEADER_FILL
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER_ALL
        col += cs
    ws.row_dimensions[row].height = 28
    row += 1

    sub_headers = [None, None, None, "с начала проведения работ", "с начала года", "в т.ч. за отчётный период"]
    for col_idx, text in enumerate(sub_headers, 1):
        if text:
            cell = ws.cell(row=row, column=col_idx, value=text)
            cell.font = _FONT_HDR
            cell.fill = _HEADER_FILL
            cell.alignment = _ALIGN_CENTER
            cell.border = _BORDER_ALL
        else:
            ws.cell(row=row, column=col_idx).border = _BORDER_ALL
            ws.cell(row=row, column=col_idx).fill = _HEADER_FILL
    # Объединяем колонки 1-3 с верхней строкой
    for c in (1, 2, 3):
        ws.merge_cells(start_row=row - 1, start_column=c, end_row=row, end_column=c)
    ws.row_dimensions[row].height = 28
    row += 1

    for col in range(1, 7):
        cell = ws.cell(row=row, column=col, value=col)
        cell.alignment = _ALIGN_CENTER
        cell.font = Font(size=8, italic=True)
        cell.border = _BORDER_ALL
    row += 1

    for idx, item in enumerate(doc.items or [], 1):
        ws.cell(row=row, column=1, value=item.position_number or idx).alignment = _ALIGN_CENTER
        ws.cell(row=row, column=2, value=item.work_name or "").alignment = _ALIGN_LEFT
        ws.cell(row=row, column=3, value=item.code or "").alignment = _ALIGN_CENTER
        for col_idx, fld in [(4, "cost_since_start"), (5, "cost_since_year"), (6, "cost_period")]:
            val = item.get(fld)
            cell = ws.cell(row=row, column=col_idx, value=flt(val) if val else None)
            cell.alignment = _ALIGN_RIGHT
            cell.number_format = '#,##0.00'
        for col in range(1, 7):
            ws.cell(row=row, column=col).font = _FONT_BODY
            ws.cell(row=row, column=col).border = _BORDER_ALL
        row += 1

    # Итого
    def _total_row(label: str, v_start, v_year, v_period) -> int:
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.cell(row=row, column=1, value=label).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=1).font = _FONT_HDR
        for col_idx, val in [(4, v_start), (5, v_year), (6, v_period)]:
            cell = ws.cell(row=row, column=col_idx, value=flt(val) if val is not None else None)
            cell.alignment = _ALIGN_RIGHT
            cell.font = _FONT_HDR
            cell.number_format = '#,##0.00'
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = _TOTAL_FILL
            ws.cell(row=row, column=col).border = _BORDER_ALL
        cur = row
        row += 1
        return cur

    _total_row("Итого:", doc.total_since_start, doc.total_since_year, doc.total_period)

    # НДС
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value=f"Сумма НДС ({doc.vat_rate or 20}%):").alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=1).font = _FONT_BODY
    ws.cell(row=row, column=6, value=flt(doc.vat_amount)).number_format = '#,##0.00'
    ws.cell(row=row, column=6).alignment = _ALIGN_RIGHT
    ws.cell(row=row, column=6).font = _FONT_BODY
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = _BORDER_ALL
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="Всего с учётом НДС:").alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=1).font = _FONT_HDR
    ws.cell(row=row, column=6, value=flt(doc.total_with_vat)).number_format = '#,##0.00'
    ws.cell(row=row, column=6).alignment = _ALIGN_RIGHT
    ws.cell(row=row, column=6).font = _FONT_HDR
    for col in range(1, 7):
        ws.cell(row=row, column=col).fill = _TOTAL_FILL
        ws.cell(row=row, column=col).border = _BORDER_ALL
    row += 1

    if doc.retention_pct and flt(doc.retention_pct) > 0:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value=f"Гарантийное удержание ({doc.retention_pct}%):").alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=1).font = _FONT_BODY
        ws.cell(row=row, column=6, value=-flt(doc.retention_amount)).number_format = '#,##0.00'
        ws.cell(row=row, column=6).alignment = _ALIGN_RIGHT
        ws.cell(row=row, column=6).font = _FONT_BODY
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = _BORDER_ALL
        row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value="Итого к оплате:").alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=1).font = _FONT_HDR
        ws.cell(row=row, column=6, value=flt(doc.total_to_pay)).number_format = '#,##0.00'
        ws.cell(row=row, column=6).alignment = _ALIGN_RIGHT
        ws.cell(row=row, column=6).font = _FONT_HDR
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = _TOTAL_FILL
            ws.cell(row=row, column=col).border = _BORDER_ALL
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Заказчик (Генподрядчик):").font = _FONT_HDR
    ws.cell(row=row, column=4, value="Подрядчик (Субподрядчик):").font = _FONT_HDR
    row += 2
    customer_line = f"{doc.signatory_customer_position or ''}  _______________  {doc.signatory_customer_name or ''}".strip()
    contractor_line = f"{doc.signatory_contractor_position or 'Директор'}  _______________  {doc.signatory_contractor_name or 'Клочков Д.А.'}"
    ws.cell(row=row, column=1, value=customer_line).font = _FONT_BODY
    ws.cell(row=row, column=4, value=contractor_line).font = _FONT_BODY
    row += 1
    ws.cell(row=row, column=1, value="должность / подпись / расшифровка").font = Font(size=8, italic=True)
    ws.cell(row=row, column=4, value="должность / подпись / расшифровка").font = Font(size=8, italic=True)
    row += 1
    ws.cell(row=row, column=1, value="М.П.").font = Font(size=9)
    ws.cell(row=row, column=4, value="М.П.").font = Font(size=9)

    _excel_response(wb, f"KS-3_{doc.name}.xlsx")


# ────────────────────────── Печать сметы ────────────────────────────────────


@frappe.whitelist()
def estimate_excel(name: str) -> None:
    """Excel-выгрузка сметы — обычная локальная форма (не гос.форма)."""
    frappe.has_permission("Estimate", "read", doc=name, throw=True)
    est = frappe.get_doc("Estimate", name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Смета"

    widths = [5, 14, 42, 8, 11, 14, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    # Шапка
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    cell = ws.cell(row=row, column=1, value=f"СМЕТА: {est.title or est.name}")
    cell.font = Font(name="Arial", size=14, bold=True)
    cell.alignment = Alignment(horizontal="center")
    row += 1

    if est.estimate_date:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        ws.cell(row=row, column=1, value=f"Дата: {formatdate(est.estimate_date, 'dd.MM.yyyy')}").font = Font(size=10, italic=True)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        row += 1

    if est.project:
        proj_title = frappe.db.get_value("Construction Project", est.project, "title") or est.project
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        ws.cell(row=row, column=1, value=f"Проект: {proj_title}").font = Font(size=10)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        row += 1

    row += 1

    # Шапка таблицы
    headers = ["№", "Код", "Наименование работ", "Ед.", "Кол-во",
               "Ед. цена (норм)", "Сумма (норм)", "Наша ед. цена", "Наша сумма"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(name="Arial", size=10, bold=True)
        c.fill = PatternFill("solid", fgColor="D8E4FC")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    ws.row_dimensions[row].height = 32
    row += 1

    idx = 0
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for item in est.items or []:
        if item.is_section:
            # Раздел — мерж по всей строке, жирный
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            c = ws.cell(row=row, column=1, value=item.section_title or item.item_name)
            c.font = Font(name="Arial", size=11, bold=True)
            c.fill = PatternFill("solid", fgColor="F0F0F0")
            c.alignment = Alignment(horizontal="left", indent=1)
            row += 1
            continue

        idx += 1
        ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=item.item_code or "").alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=item.item_name or "").alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=4, value=item.unit or "").alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=flt(item.qty) if item.qty else None).number_format = '#,##0.000'
        ws.cell(row=row, column=6, value=flt(item.base_unit_price) if item.base_unit_price else None).number_format = '#,##0.00'
        ws.cell(row=row, column=7, value=flt(item.base_amount) if item.base_amount else None).number_format = '#,##0.00'
        ws.cell(row=row, column=8, value=flt(item.our_unit_price) if item.our_unit_price else None).number_format = '#,##0.00'
        ws.cell(row=row, column=9, value=flt(item.our_amount) if item.our_amount else None).number_format = '#,##0.00'

        for col in range(1, 10):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).font = Font(name="Arial", size=10)
        row += 1

    # Итоги
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value="Итого по нормам:").alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=row, column=7, value=flt(est.base_total)).number_format = '#,##0.00'
    ws.cell(row=row, column=7).font = Font(bold=True, size=11)
    ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value="Наша цена (с маржой):").alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=row, column=9, value=flt(est.our_total)).number_format = '#,##0.00'
    ws.cell(row=row, column=9).font = Font(bold=True, size=12, color="C8460F")
    ws.cell(row=row, column=9).alignment = Alignment(horizontal="right")
    row += 1

    if flt(est.margin_pct):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row=row, column=1, value=f"Маржа: {flt(est.margin_pct):.2f}%").alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=1).font = Font(size=10, italic=True, color="666666")

    _excel_response(wb, f"Smeta_{est.name}.xlsx")


_ESTIMATE_PDF_TEMPLATE = """<!DOCTYPE html><html lang=ru><head><meta charset=utf-8>
<title>Смета {{ doc.name }}</title>
<style>
@page { size: A4 landscape; margin: 10mm; }
body { font-family: "DejaVu Sans", Arial, sans-serif; font-size: 9pt; color: #000; }
h1 { text-align: center; font-size: 14pt; margin: 0 0 4px; }
.meta { text-align: center; font-size: 9pt; color: #333; margin-bottom: 14px; }
table { width: 100%; border-collapse: collapse; }
th, td { border: 1px solid #000; padding: 4px 6px; vertical-align: top; }
th { background: #d8e4fc; font-weight: 700; text-align: center; font-size: 9pt; }
td.num { text-align: right; font-variant-numeric: tabular-nums; white-space: nowrap; }
td.c { text-align: center; }
tr.section td { background: #f0f0f0; font-weight: 700; }
.totals { margin-top: 16px; }
.totals td { border: none; padding: 3px 6px; }
.totals .label { text-align: right; font-weight: 700; }
.totals .value { text-align: right; font-family: monospace; font-weight: 700; }
.totals .our { color: #c8460f; font-size: 12pt; }
</style></head>
<body>
<h1>СМЕТА: {{ doc.title or doc.name }}</h1>
<div class="meta">
  {% if doc.estimate_date %}{{ frappe.utils.formatdate(doc.estimate_date, "dd.MM.yyyy") }}{% endif %}
  {% if project_title %} · Проект: {{ project_title }}{% endif %}
  {% if doc.estimation_method %} · Метод: {{ doc.estimation_method }}{% endif %}
</div>
<table>
<thead><tr>
  <th style="width:4%">№</th><th style="width:10%">Код</th>
  <th style="width:36%">Наименование работ</th><th style="width:6%">Ед.</th>
  <th style="width:8%">Кол-во</th>
  <th style="width:9%">Ед.цена (норм)</th><th style="width:9%">Сумма (норм)</th>
  <th style="width:9%">Наша ед.</th><th style="width:9%">Наша сумма</th>
</tr></thead>
<tbody>
{% set ns = namespace(idx=0) %}
{% for item in (doc.items or []) %}
  {% if item.is_section %}
    <tr class="section"><td colspan="9">{{ item.section_title or item.item_name }}</td></tr>
  {% else %}
    {% set ns.idx = ns.idx + 1 %}
    <tr>
      <td class="c">{{ ns.idx }}</td>
      <td class="c">{{ item.item_code or "" }}</td>
      <td>{{ item.item_name or "" }}</td>
      <td class="c">{{ item.unit or "" }}</td>
      <td class="num">{{ "{:,.3f}".format(item.qty|float).replace(",", " ").replace(".", ",") if item.qty else "—" }}</td>
      <td class="num">{{ "{:,.2f}".format(item.base_unit_price|float).replace(",", " ").replace(".", ",") if item.base_unit_price else "—" }}</td>
      <td class="num">{{ "{:,.2f}".format(item.base_amount|float).replace(",", " ").replace(".", ",") if item.base_amount else "—" }}</td>
      <td class="num">{{ "{:,.2f}".format(item.our_unit_price|float).replace(",", " ").replace(".", ",") if item.our_unit_price else "—" }}</td>
      <td class="num">{{ "{:,.2f}".format(item.our_amount|float).replace(",", " ").replace(".", ",") if item.our_amount else "—" }}</td>
    </tr>
  {% endif %}
{% endfor %}
</tbody>
</table>
<table class="totals">
  <tr><td class="label" colspan="6">Итого по нормам:</td><td class="value">{{ "{:,.2f}".format(doc.base_total|float).replace(",", " ").replace(".", ",") if doc.base_total else "—" }} ₽</td></tr>
  <tr><td class="label our" colspan="8">Наша цена:</td><td class="value our">{{ "{:,.2f}".format(doc.our_total|float).replace(",", " ").replace(".", ",") if doc.our_total else "—" }} ₽</td></tr>
  {% if doc.margin_pct %}<tr><td colspan="9" style="text-align:right;color:#666;font-style:italic">Маржа: {{ "{:.2f}".format(doc.margin_pct|float) }}%</td></tr>{% endif %}
</table>
</body></html>
"""


@frappe.whitelist()
def estimate_pdf(name: str) -> None:
    """PDF сметы (через wkhtmltopdf + локальная Jinja-шаблон, A4 альбомная)."""
    frappe.has_permission("Estimate", "read", doc=name, throw=True)
    doc = frappe.get_doc("Estimate", name)

    project_title = ""
    if doc.project and frappe.db.exists("Construction Project", doc.project):
        project_title = frappe.db.get_value("Construction Project", doc.project, "title") or doc.project

    html = frappe.render_template(_ESTIMATE_PDF_TEMPLATE, {
        "doc": doc,
        "project_title": project_title,
        "frappe": frappe,
    })

    pdf_bytes = get_pdf(html, options={
        "page-size": "A4",
        "orientation": "Landscape",
        "margin-top": "10mm",
        "margin-bottom": "10mm",
        "margin-left": "10mm",
        "margin-right": "10mm",
        "encoding": "UTF-8",
    })

    frappe.local.response.filename = f"Smeta_{doc.name}.pdf"
    frappe.local.response.filecontent = pdf_bytes
    frappe.local.response.type = "pdf"


# ────────────────────────── Excel-экспорт списков ────────────────────────────


_LIST_SPECS = {
    "tenders": {
        "doctype": "Tender",
        "sheet": "Тендеры",
        "filename": "Tenders",
        "columns": [
            ("name", "Код", 15),
            ("title", "Название", 50),
            ("status", "Статус", 18),
            ("tender_law", "Закупка", 14),
            ("work_type", "Тип работ", 16),
            ("region", "Регион", 18),
            ("nmck", "НМЦК, ₽", 14),
            ("our_price", "Наша цена, ₽", 14),
            ("margin_pct", "Маржа, %", 11),
            ("deadline_date", "Дедлайн", 12),
            ("ai_match_score", "AI score", 10),
        ],
    },
    "projects": {
        "doctype": "Construction Project",
        "sheet": "Проекты",
        "filename": "Projects",
        "columns": [
            ("name", "Код", 14),
            ("title", "Название", 40),
            ("status", "Статус", 14),
            ("customer", "Заказчик", 24),
            ("work_type", "Тип работ", 16),
            ("location", "Адрес", 30),
            ("contract_amount", "Контракт, ₽", 14),
            ("planned_cost", "План.себест., ₽", 16),
            ("real_revenue", "Выручка, ₽", 14),
            ("real_cost", "Расходы, ₽", 14),
            ("real_margin_pct", "Маржа факт, %", 13),
            ("ks2_completion_pct", "Закрыто КС-2, %", 14),
            ("start_date", "Начало", 12),
            ("planned_end_date", "Сдача план", 12),
        ],
    },
    "estimates": {
        "doctype": "Estimate",
        "sheet": "Сметы",
        "filename": "Estimates",
        "columns": [
            ("name", "Код", 14),
            ("title", "Название", 40),
            ("status", "Статус", 16),
            ("project", "Проект", 14),
            ("tender", "Тендер", 14),
            ("estimate_date", "Дата", 12),
            ("base_total", "По нормам, ₽", 14),
            ("our_total", "Наша цена, ₽", 14),
            ("margin_pct", "Маржа, %", 11),
            ("estimation_method", "Метод оценки", 22),
        ],
    },
    "stock": {
        "doctype": "Stock Item",
        "sheet": "Остатки",
        "filename": "Stock",
        "columns": [
            ("name", "ID", 14),
            ("item_name", "Материал", 40),
            ("item_code", "Артикул", 14),
            ("category", "Категория", 20),
            ("unit", "Ед.", 7),
            ("default_warehouse", "Склад", 18),
            ("current_qty", "Остаток", 12),
            ("min_qty", "Мин.", 10),
            ("avg_price", "Ср. цена, ₽", 13),
            ("total_value", "Стоимость, ₽", 14),
            ("last_movement_date", "Посл. движение", 14),
        ],
    },
    "certifications": {
        "doctype": "Employee Certification",
        "sheet": "Аттестации",
        "filename": "Certifications",
        "columns": [
            ("name", "Код", 14),
            ("employee_name", "ФИО", 26),
            ("employee_role", "Должность", 20),
            ("cert_type", "Тип", 32),
            ("cert_number", "№", 18),
            ("issuing_organization", "Кем выдано", 24),
            ("issue_date", "Выдано", 12),
            ("expiry_date", "Действует до", 13),
            ("status", "Статус", 14),
        ],
    },
    "ks2": {
        "doctype": "KS2 Act",
        "sheet": "КС-2",
        "filename": "KS2",
        "columns": [
            ("name", "Код", 14),
            ("title", "Название", 36),
            ("status", "Статус", 14),
            ("project", "Проект", 14),
            ("customer", "Заказчик", 24),
            ("contract_number", "№ договора", 16),
            ("act_date", "Дата акта", 12),
            ("period_from", "Период с", 12),
            ("period_to", "Период по", 12),
            ("amount", "Сумма, ₽", 14),
            ("payment_status", "Оплата", 12),
            ("payment_received", "Получено, ₽", 14),
        ],
    },
}


@frappe.whitelist()
def export_list(spec: str) -> None:
    """Универсальный Excel-экспорт списка одного из поддерживаемых DocType.

    spec ∈ ('tenders','projects','estimates','stock','certifications','ks2').
    """
    if spec not in _LIST_SPECS:
        frappe.throw(_(f"Неизвестный экспорт: {spec}. Доступно: {list(_LIST_SPECS.keys())}"))

    cfg = _LIST_SPECS[spec]
    dt = cfg["doctype"]
    frappe.has_permission(dt, "read", throw=True)

    fields = [c[0] for c in cfg["columns"]]
    rows = frappe.get_all(dt, fields=fields, order_by="modified desc", limit=5000)

    wb = Workbook()
    ws = wb.active
    ws.title = cfg["sheet"]

    # Ширины
    for i, (_fld, _label, w) in enumerate(cfg["columns"], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Заголовки
    header_font = Font(name="Arial", size=10, bold=True)
    header_fill = PatternFill("solid", fgColor="D8E4FC")
    for i, (_fld, label, _w) in enumerate(cfg["columns"], 1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDER_ALL
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    body_font = Font(name="Arial", size=10)
    money_fields = {"nmck", "our_price", "contract_amount", "planned_cost", "real_revenue",
                    "real_cost", "base_total", "our_total", "amount", "payment_received",
                    "avg_price", "total_value"}
    pct_fields = {"margin_pct", "real_margin_pct", "ks2_completion_pct"}
    date_fields = {"deadline_date", "start_date", "planned_end_date", "estimate_date",
                   "issue_date", "expiry_date", "last_movement_date", "act_date",
                   "period_from", "period_to"}

    for row_idx, row in enumerate(rows, 2):
        for col_idx, (fld, _label, _w) in enumerate(cfg["columns"], 1):
            val = row.get(fld)
            cell = ws.cell(row=row_idx, column=col_idx)
            if val is None or val == "":
                cell.value = ""
            elif fld in money_fields:
                cell.value = flt(val)
                cell.number_format = '#,##0.00'
            elif fld in pct_fields:
                cell.value = flt(val)
                cell.number_format = '0.00"%"'
            elif fld in date_fields:
                cell.value = str(val)
                cell.alignment = Alignment(horizontal="center")
            elif fld in ("current_qty", "min_qty"):
                cell.value = flt(val)
                cell.number_format = '#,##0.000'
            else:
                cell.value = str(val)
            cell.font = body_font
            cell.border = _BORDER_ALL
            if fld in money_fields or fld in pct_fields or fld in ("current_qty", "min_qty"):
                cell.alignment = Alignment(horizontal="right")

    _excel_response(wb, f"{cfg['filename']}_{frappe.utils.nowdate()}.xlsx")
