"""API для акта сверки расчётов с контрагентом.

Идея (1С / СБИС / Контур): формируем стандартный акт сверки за период:
- Начальное сальдо
- Дебет (что мы выполнили — КС-2 + Change Order amount_approved)
- Кредит (что заказчик оплатил — payment_received из КС-2 + Material Request для поставщиков)
- Конечное сальдо

Минимальная версия v1: Customer (заказчик) — выполнение через КС-2, оплата через payment_received.
Расширение для Supplier (Material Request) — в следующей итерации.
"""
from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO

import frappe
from frappe.utils import flt, getdate


@frappe.whitelist()
def list_partners(party_type: str = "Customer", days: int = 365) -> list[dict]:
    """Контрагенты с активностью за период (для выбора в выпадашке)."""
    frappe.has_permission("Customer", throw=True)

    if party_type == "Customer":
        rows = frappe.db.sql("""
            SELECT k.customer AS party,
                   COUNT(*) AS ks2_count,
                   SUM(k.amount) AS total_amount,
                   MAX(k.act_date) AS last_act
            FROM `tabKS2 Act` k
            WHERE k.customer IS NOT NULL
              AND k.act_date >= DATE_SUB(CURDATE(), INTERVAL %(d)s DAY)
            GROUP BY k.customer
            ORDER BY total_amount DESC
        """, {"d": int(days)}, as_dict=True)

        for r in rows:
            r["party_name"] = frappe.db.get_value("Customer", r["party"], "customer_name") or r["party"]
            r["total_amount"] = float(r["total_amount"] or 0)
            r["last_act"] = str(r["last_act"]) if r["last_act"] else None
        return rows

    frappe.throw(f"party_type {party_type} пока не поддерживается")


@frappe.whitelist()
def build_reconciliation(party_type: str, party: str,
                         from_date: str | None = None,
                         to_date: str | None = None) -> dict:
    """Строит акт сверки за период.

    Структура возврата:
    {
      "party": ..., "party_name": ...,
      "from_date": ..., "to_date": ...,
      "opening_balance": float,  # сальдо на начало (наш долг заказчику = плюс, его долг нам = минус)
      "rows": [
        {"date", "doc_type", "doc_number", "description", "debit", "credit", "balance"}
      ],
      "totals": {"debit": .., "credit": .., "closing_balance": ..},
    }
    Логика для Customer:
    - debit (мы делали): подписанные КС-2.amount, одобренные Change Order
    - credit (мы получили): KS2.payment_received где payment_status указывает оплачено
    """
    frappe.has_permission("Customer", throw=True)

    if party_type != "Customer":
        frappe.throw("Поддерживается пока только Customer")
    if not party or not frappe.db.exists("Customer", party):
        frappe.throw(f"Customer {party} не найден")

    today = getdate()
    to_d = getdate(to_date) if to_date else today
    from_d = getdate(from_date) if from_date else (to_d - timedelta(days=90))

    party_name = frappe.db.get_value("Customer", party, "customer_name") or party

    # ── Opening balance: все события ДО from_date ──────────────────────────
    open_ks2 = frappe.db.sql("""
        SELECT COALESCE(SUM(amount), 0) AS debit,
               COALESCE(SUM(CASE WHEN payment_status IN ('Оплачена','Получено') OR payment_received=1
                                THEN amount ELSE 0 END), 0) AS credit
        FROM `tabKS2 Act`
        WHERE customer = %(p)s AND act_date < %(d)s AND status = 'Подписан'
    """, {"p": party, "d": str(from_d)}, as_dict=True)[0]

    opening = float(open_ks2["debit"] or 0) - float(open_ks2["credit"] or 0)

    # ── Строки за период ────────────────────────────────────────────────────
    rows: list[dict] = []

    # КС-2 подписанные в периоде
    ks2_period = frappe.db.sql("""
        SELECT name, act_number, act_date, amount, payment_status, payment_received,
               period_from, period_to, signed_date
        FROM `tabKS2 Act`
        WHERE customer = %(p)s AND status = 'Подписан'
              AND act_date BETWEEN %(s)s AND %(e)s
        ORDER BY act_date ASC, name ASC
    """, {"p": party, "s": str(from_d), "e": str(to_d)}, as_dict=True)

    # Change Order одобренные в периоде
    has_change_order = frappe.db.exists("DocType", "Change Order")
    co_period: list[dict] = []
    if has_change_order:
        try:
            co_period = frappe.db.sql("""
                SELECT co.name, co.title, co.amount_approved, co.approved_date
                FROM `tabChange Order` co
                INNER JOIN `tabConstruction Project` p ON p.name = co.project
                WHERE p.customer = %(p)s
                      AND co.status = 'Одобрено'
                      AND co.approved_date BETWEEN %(s)s AND %(e)s
                ORDER BY co.approved_date ASC
            """, {"p": party, "s": str(from_d), "e": str(to_d)}, as_dict=True)
        except Exception:
            co_period = []

    # Merge + sort by date
    events: list[dict] = []
    for k in ks2_period:
        events.append({
            "date": k["act_date"],
            "doc_type": "КС-2",
            "doc_number": k["act_number"] or k["name"],
            "description": f"Выполненные работы (период {k['period_from']}–{k['period_to']})" if k.get("period_from") else "Выполненные работы",
            "debit": float(k["amount"] or 0),
            "credit": 0.0,
            "_ks2_name": k["name"],
            "_payment_status": k.get("payment_status"),
            "_payment_received": k.get("payment_received"),
        })
        # Если КС-2 оплачен — добавляем credit-строку (виртуальная)
        if (k.get("payment_status") in ("Оплачена", "Получено")
                or k.get("payment_received") == 1):
            events.append({
                "date": k.get("signed_date") or k["act_date"],
                "doc_type": "Оплата",
                "doc_number": f"к {k['act_number'] or k['name']}",
                "description": "Поступление оплаты от заказчика",
                "debit": 0.0,
                "credit": float(k["amount"] or 0),
            })

    for c in co_period:
        events.append({
            "date": c["approved_date"],
            "doc_type": "Доп. соглашение",
            "doc_number": c["name"],
            "description": c["title"] or "Изменения объёмов",
            "debit": float(c["amount_approved"] or 0),
            "credit": 0.0,
        })

    events.sort(key=lambda x: (x["date"] or date.min, x.get("doc_type") == "Оплата"))

    balance = opening
    total_debit = total_credit = 0.0
    for e in events:
        balance += float(e["debit"]) - float(e["credit"])
        total_debit += float(e["debit"])
        total_credit += float(e["credit"])
        rows.append({
            "date": str(e["date"]),
            "doc_type": e["doc_type"],
            "doc_number": e["doc_number"],
            "description": e["description"],
            "debit": e["debit"],
            "credit": e["credit"],
            "balance": balance,
        })

    return {
        "party": party,
        "party_name": party_name,
        "party_type": party_type,
        "from_date": str(from_d),
        "to_date": str(to_d),
        "opening_balance": opening,
        "rows": rows,
        "totals": {
            "debit": total_debit,
            "credit": total_credit,
            "closing_balance": balance,
        },
        "as_of": str(today),
    }


@frappe.whitelist(allow_guest=False, methods=["GET"])
def export_xlsx(party_type: str, party: str,
                from_date: str | None = None,
                to_date: str | None = None) -> None:
    """Excel-выгрузка акта сверки. Возвращает файл."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    rec = build_reconciliation(party_type, party, from_date, to_date)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Акт сверки"

    # Header block
    ws["A1"] = "Акт сверки взаимных расчётов"
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells("A1:F1")

    ws["A3"] = f"Контрагент: {rec['party_name']}"
    ws["A3"].font = Font(bold=True)
    ws["A4"] = f"Период: с {rec['from_date']} по {rec['to_date']}"
    ws["A5"] = f"Сальдо на начало периода: {rec['opening_balance']:,.2f} ₽".replace(",", " ")

    # Table header
    header_row = 7
    ws.cell(row=header_row, column=1, value="Дата")
    ws.cell(row=header_row, column=2, value="Тип")
    ws.cell(row=header_row, column=3, value="№ документа")
    ws.cell(row=header_row, column=4, value="Содержание")
    ws.cell(row=header_row, column=5, value="Дебет (₽)")
    ws.cell(row=header_row, column=6, value="Кредит (₽)")
    ws.cell(row=header_row, column=7, value="Сальдо (₽)")

    fill = PatternFill("solid", fgColor="E8E8E8")
    bold = Font(bold=True)
    thin = Side(border_style="thin", color="999999")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for col in range(1, 8):
        c = ws.cell(row=header_row, column=col)
        c.font = bold
        c.fill = fill
        c.border = border

    # Data rows
    r = header_row + 1
    for row in rec["rows"]:
        ws.cell(row=r, column=1, value=row["date"])
        ws.cell(row=r, column=2, value=row["doc_type"])
        ws.cell(row=r, column=3, value=row["doc_number"])
        ws.cell(row=r, column=4, value=row["description"])
        ws.cell(row=r, column=5, value=row["debit"] or None)
        ws.cell(row=r, column=6, value=row["credit"] or None)
        ws.cell(row=r, column=7, value=row["balance"])
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = border
        # Числовой формат
        for col in (5, 6, 7):
            ws.cell(row=r, column=col).number_format = "#,##0.00"
        r += 1

    # Totals
    r += 1
    ws.cell(row=r, column=1, value="ИТОГО за период")
    ws.cell(row=r, column=1).font = bold
    ws.cell(row=r, column=5, value=rec["totals"]["debit"]).font = bold
    ws.cell(row=r, column=6, value=rec["totals"]["credit"]).font = bold
    ws.cell(row=r, column=7, value=rec["totals"]["closing_balance"]).font = bold
    for col in (5, 6, 7):
        ws.cell(row=r, column=col).number_format = "#,##0.00"

    r += 2
    closing = rec["totals"]["closing_balance"]
    if closing > 0.01:
        ws.cell(row=r, column=1, value=f"Задолженность {rec['party_name']} перед нами: {closing:,.2f} ₽".replace(",", " "))
    elif closing < -0.01:
        ws.cell(row=r, column=1, value=f"Наша задолженность перед {rec['party_name']}: {-closing:,.2f} ₽".replace(",", " "))
    else:
        ws.cell(row=r, column=1, value="Расчёты закрыты, задолженности нет.")
    ws.cell(row=r, column=1).font = Font(bold=True, size=12)

    # Подписи
    r += 3
    ws.cell(row=r, column=1, value="От ООО «Олимп»: ___________________ /_____________/")
    ws.cell(row=r + 1, column=1, value=f"От {rec['party_name']}: ___________________ /_____________/")

    # Ширины
    widths = {1: 13, 2: 18, 3: 22, 4: 50, 5: 16, 6: 16, 7: 18}
    for col, w in widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    for col in range(1, 8):
        for row in range(header_row, r + 1):
            ws.cell(row=row, column=col).alignment = Alignment(vertical="center", wrap_text=col == 4)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_party = rec["party_name"][:30].replace(" ", "_").replace("/", "-")
    filename = f"recon_{safe_party}_{rec['from_date']}_{rec['to_date']}.xlsx"

    frappe.local.response.filename = filename
    frappe.local.response.filecontent = buf.getvalue()
    frappe.local.response.type = "download"
